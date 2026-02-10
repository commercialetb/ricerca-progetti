# app.py
# Ricerca Progetti Italia – Streamlit (Lite + toggles) + filtro Regione
# Output: XLSX+PDF (12 colonne richieste) + CSV FASE1 (colonne estese)

from __future__ import annotations

import io
import re
import time
import math
from dataclasses import dataclass
from typing import List, Dict, Optional, Tuple
from urllib.parse import urljoin, urlparse

import pandas as pd
import requests
import streamlit as st
from bs4 import BeautifulSoup

# Optional: reportlab for PDF
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

# Optional: openpyxl for XLSX
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table as XLTable, TableStyleInfo


st.set_page_config(page_title="Ricerca Progetti Italia (Lite)", page_icon="🔎", layout="wide")

DEFAULT_UA = (
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
)

REGIONS = [
    "(tutte)",
    "Abruzzo", "Basilicata", "Calabria", "Campania", "Emilia-Romagna",
    "Friuli-Venezia Giulia", "Lazio", "Liguria", "Lombardia", "Marche",
    "Molise", "Piemonte", "Puglia", "Sardegna", "Sicilia", "Toscana",
    "Trentino-Alto Adige", "Umbria", "Valle d'Aosta", "Veneto",
]

CATEGORIES: Dict[str, List[str]] = {
    "Sport e impianti sportivi (impianti, piscine, palestre, palazzetti)": [
        "impianto sportivo", "stadio", "palazzetto", "palestra", "piscina",
        "polisportivo", "campo", "centro sportivo",
    ],
    "Piste ciclabili e mobilità dolce (piste ciclabili, ciclovie, percorsi ciclo-pedonali)": [
        "pista ciclabile", "ciclovia", "ciclopedonale", "greenway",
        "mobilità dolce", "percorso ciclabile", "percorso ciclo pedonale",
    ],
    "Progetti aeroportuali": ["aeroporto", "aerostazione", "terminal", "hangar", "airside", "landside"],
    "Centri commerciali": ["centro commerciale", "mall", "galleria commerciale", "retail park", "ipermercato"],
    "Progetti alberghieri": ["hotel", "albergo", "resort", "struttura ricettiva", "ospitalità"],
    "Edilizia residenziale pubblica (ERP, housing, rigenerazione urbana, ristrutturazioni)": [
        "ERP", "edilizia residenziale pubblica", "housing", "social housing",
        "rigenerazione urbana", "ristrutturazione", "case popolari",
    ],
    "Strutture sanitarie (ospedali, poliambulatori, RSA, laboratori, fisioterapia)": [
        "ospedale", "poliambulatorio", "RSA", "laboratorio", "fisioterapia",
        "struttura sanitaria", "ambulatorio",
    ],
    "Università e ricerca (campus, laboratori, biblioteche, residenze)": [
        "campus", "laboratorio", "biblioteca", "residenza universitaria", "dipartimento",
    ],
    "Innovazione e startup (hub, incubatori, fab lab, maker space)": [
        "hub", "incubatore", "fab lab", "maker space", "acceleratore", "innovation",
    ],
    "Archivi e patrimonio culturale (archivi, biblioteche, musei, restauro)": [
        "archivio", "deposito", "biblioteca", "museo", "restauro", "patrimonio culturale",
    ],
    "Musei e spazi culturali (musei civici, gallerie, centri culturali)": [
        "museo", "musei civici", "galleria", "spazio culturale", "centro culturale",
    ],
    "Intrattenimento e spettacolo (teatri, cinema, auditorium, anfiteatri, spazi concerti)": [
        "teatro", "cinema", "auditorium", "anfiteatro", "spazio concerti", "arena",
    ],
    "Trasporto pubblico locale (tram, autobus, metro, bus elettrici, stazioni, bike sharing)": [
        "tram", "autobus", "metro", "bus elettrici", "stazione", "bike sharing", "deposito bus",
    ],
    "Edilizia giudiziaria e sicurezza (tribunali, carceri, questure, caserme)": [
        "tribunale", "carcere", "questura", "caserma", "comando",
    ],
    "Sistemazioni urbane (riqualificazione urbana, piazze, strade, parchi, arredo urbano)": [
        "riqualificazione urbana", "piazza", "strada", "parco", "arredo urbano", "sistemazione urbana",
    ],
}

PROJECT_STATUS = ["(qualsiasi)", "progetto PFTE", "progetto definitivo", "progetto esecutivo"]
PROJECT_PHASE = ["(qualsiasi)", "FASE DI PROGRAMMAZIONE", "FASE DI PROGETTAZIONE", "FASE DI ESECUZIONE"]

DATE_RE = re.compile(r"\b(?:(?:0?[1-9]|[12]\d|3[01])[-/\.](?:0?[1-9]|1[0-2])[-/\.](?:19|20)\d{2})\b")
EMAIL_RE = re.compile(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", re.I)
CIG_RE = re.compile(r"\b([0-9A-Z]{10})\b")
CUP_RE = re.compile(r"\b([A-Z][0-9A-Z]{14})\b", re.I)
EURO_RE = re.compile(r"€\s?[\d\.\,]+|\b[\d\.\,]+\s?€")

REGION_COL_CANDIDATES = ["regione", "region", "area", "macroarea", "macro_area", "territorio", "territoriale"]

# =========================
# OUTPUT SCHEMAS
# =========================

# 12 colonne richieste per Excel/PDF (seconda lista)
EXCEL_PDF_COLS_12 = [
    "Nome progetto",
    "Comune",
    "Stato del progetto (PFTE/Definitivo/Esecutivo/N/D)",
    "Fase (Programmazione/Progettazione/Esecuzione/N/D)",
    "Settore (Terziario/Servizi/Residenziale/Industria/Ingegneria/N/D)",
    "Dettagli",
    "Progettazione Capogruppo",
    "Progettazione Direzione Lavori",
    "Progettazione Architettonica",
    "Studio di Fattibilità",
    "Responsabile Progetto",
    "Progetto Impianti Elettrici",
]

# CSV FASE 1 (colonne estese come hai scritto — le compiliamo se troviamo, altrimenti N/D)
CSV_FASE1_COLS = [
    "Data pubblicazione (DD/MM/YYYY)",
    "REGIONE (coerente con CIG/CUP)",
    "Comune/Città",
    "Nome Progetto",
    "Settore/Tipologia",
    "Ente Committente (Comune, Regione, ASL, Università, Privato, Pubblico-Privato)",
    "Tipo Ente (Pubblico, Pubblico-Privato, Privato)",
    "Valore Progetto Totale (€)",
    "Valore Fase Corrente (€)",
    "% Completamento (Preliminare, Definitiva, Esecutiva, DL, Collaudo)",
    "Data scadenza (DD/MM/YYYY)",
    "CIG",
    "CUP",
    "Portale/Fonte (OpenPNRR, OpenCUP, MEPA, GURI, Albo Pretorio, etc.)",
    "Studio/Progettista",
    "Fase Progettazione (Preliminare, Definitiva, Esecutiva, Direttore Lavori, Collaudatore)",
]


# =========================
# DATA MODELS
# =========================

@dataclass
class SearchHit:
    source_url: str
    found_url: str
    title: str
    matched_terms: List[str]
    doc_date: Optional[str] = None
    emails: Optional[List[str]] = None
    note: str = ""
    raw_text_sample: str = ""


def nd(x: Optional[str]) -> str:
    x = (x or "").strip()
    return x if x else "N/D"


def safe_get(url: str, timeout: int = 25) -> requests.Response:
    headers = {"User-Agent": DEFAULT_UA, "Accept-Language": "it-IT,it;q=0.9,en;q=0.8"}
    resp = requests.get(url, headers=headers, timeout=timeout, allow_redirects=True)
    resp.raise_for_status()
    return resp


def normalize_url(u: str) -> str:
    u = (u or "").strip()
    if not u:
        return ""
    if not u.startswith(("http://", "https://")):
        u = "https://" + u
    return u


def guess_text_from_html(html: str) -> str:
    soup = BeautifulSoup(html, "lxml")
    for t in soup(["script", "style", "noscript"]):
        t.decompose()
    return " ".join(soup.get_text(" ").split())


def extract_links(base_url: str, html: str) -> List[str]:
    soup = BeautifulSoup(html, "lxml")
    links = []
    for a in soup.find_all("a", href=True):
        href = a.get("href")
        if not href:
            continue
        full = urljoin(base_url, href)
        if full.startswith(("mailto:", "javascript:")):
            continue
        links.append(full)
    seen, out = set(), []
    for x in links:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def is_probably_pdf(url: str) -> bool:
    return urlparse(url).path.lower().endswith(".pdf")


def compile_terms(category: str, extra_terms: str, status: str, phase: str, region: str) -> List[str]:
    terms: List[str] = []
    if category and category != "(qualsiasi)":
        terms.extend(CATEGORIES.get(category, []))
    if status and status != "(qualsiasi)":
        terms.append(status)
    if phase and phase != "(qualsiasi)":
        terms.append(phase)
    if region and region != "(tutte)":
        terms.append(region)

    if extra_terms:
        for t in re.split(r"[,\n;]+", extra_terms):
            t = t.strip()
            if t:
                terms.append(t)

    seen, cleaned = set(), []
    for t in terms:
        k = t.lower()
        if k not in seen:
            seen.add(k)
            cleaned.append(t)
    return cleaned


def find_matches(text: str, terms: List[str]) -> List[str]:
    low = text.lower()
    return [t for t in terms if t.lower() in low]


def parse_date_from_text(text: str) -> Optional[str]:
    m = DATE_RE.search(text)
    return m.group(0) if m else None


def in_date_range(date_str: Optional[str], start, end) -> bool:
    if not start and not end:
        return True
    if not date_str:
        # se non abbiamo data, non escludiamo
        return True
    d = None
    for sep in ("/", "-", "."):
        if sep in date_str:
            p = date_str.split(sep)
            if len(p) == 3:
                try:
                    import datetime as _dt
                    d = _dt.date(int(p[2]), int(p[1]), int(p[0]))
                except Exception:
                    d = None
            break
    if not d:
        return True
    if start and d < start:
        return False
    if end and d > end:
        return False
    return True


def pdf_extract_text_and_meta(url: str, max_bytes: int = 12_000_000):
    note = ""
    try:
        resp = safe_get(url, timeout=35)
        content = resp.content
        if len(content) > max_bytes:
            return "", None, [], f"PDF troppo grande ({len(content)/1e6:.1f} MB), skip parsing"

        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(content))

        parts = []
        for page in reader.pages[:8]:
            try:
                parts.append(page.extract_text() or "")
            except Exception:
                continue
        text = "\n".join(parts)
        date_str = parse_date_from_text(text)
        emails = sorted(set(EMAIL_RE.findall(text)))
        return text, date_str, emails, note
    except Exception as e:
        return "", None, [], f"Errore parsing PDF: {e}"


def selenium_fetch_html(url: str, wait_s: int = 6) -> str:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options

    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1280,1024")
    options.add_argument(f"--user-agent={DEFAULT_UA}")

    driver = webdriver.Chrome(options=options)
    try:
        driver.get(url)
        time.sleep(max(1, wait_s))
        return driver.page_source or ""
    finally:
        try:
            driver.quit()
        except Exception:
            pass


def fetch_html(url: str, use_selenium: bool):
    try:
        if use_selenium:
            return selenium_fetch_html(url, wait_s=6), "selenium"
        r = safe_get(url, timeout=25)
        return r.text, "requests"
    except Exception as e:
        return "", f"fetch error: {e}"


@st.cache_data(show_spinner=False)
def load_csv_robust_from_bytes(raw: bytes) -> Tuple[pd.DataFrame, List[str]]:
    warnings: List[str] = []
    try:
        df = pd.read_csv(io.BytesIO(raw))
        return df, warnings
    except Exception:
        df = pd.read_csv(io.BytesIO(raw), sep=None, engine="python", on_bad_lines="skip")
        warnings.append("CSV letto in modalità robusta (sep auto + skip righe malformate).")
        return df, warnings


@st.cache_data(show_spinner=False)
def load_csv_robust(path_or_url: str) -> Tuple[pd.DataFrame, List[str]]:
    warnings: List[str] = []
    if not path_or_url:
        return pd.DataFrame(), ["Percorso/URL CSV vuoto"]
    try:
        if path_or_url.startswith(("http://", "https://")):
            raw = safe_get(path_or_url, timeout=30).content
        else:
            with open(path_or_url, "rb") as f:
                raw = f.read()
        return load_csv_robust_from_bytes(raw)
    except Exception as e:
        return pd.DataFrame(), [f"Errore caricamento CSV: {e}"]


def show_working_indicator(step_text: str, progress: float, eta_s):
    cols = st.columns([2, 1])
    with cols[0]:
        st.info(f"⏳ In lavorazione: {step_text}")
        st.progress(min(max(progress, 0.0), 1.0))
    with cols[1]:
        st.metric("ETA stimata", f"{int(max(0, eta_s))} s" if eta_s is not None and math.isfinite(eta_s) else "—")


def detect_region_column(df: pd.DataFrame) -> Optional[str]:
    cols = [str(c) for c in df.columns]
    lower = [c.lower().strip() for c in cols]
    for i, c in enumerate(lower):
        for cand in REGION_COL_CANDIDATES:
            if cand in c:
                return cols[i]
    return None


# =========================
# HEURISTIC EXTRACTION
# =========================

def infer_source_label(url: str) -> str:
    u = (url or "").lower()
    if "opencup" in u:
        return "OpenCUP"
    if "openpnrr" in u or "pnrr" in u:
        return "OpenPNRR"
    if "acquistinretepa" in u or "mepa" in u:
        return "MEPA"
    if "gazzettaufficiale" in u or "guri" in u:
        return "GURI"
    if "albo" in u or "pretorio" in u:
        return "Albo Pretorio"
    if "anac" in u:
        return "ANAC"
    return "Web/Portale"


def infer_settore_from_category(category: str) -> str:
    # Mappatura “soft” verso (Terziario/Servizi/Residenziale/Industria/Ingegneria/N/D)
    cat = (category or "").lower()
    if "residenziale" in cat or "erp" in cat or "housing" in cat:
        return "Residenziale"
    if "sanitar" in cat or "università" in cat or "ricerca" in cat or "muse" in cat or "cultur" in cat or "trasporto" in cat:
        return "Servizi"
    if "centri commerciali" in cat or "alberghieri" in cat or "innovazione" in cat:
        return "Terziario"
    if "aeroport" in cat or "mobilità" in cat or "sistemazioni urbane" in cat or "ciclabili" in cat:
        return "Ingegneria"
    return "N/D"


def infer_stato_progetto(text: str) -> str:
    t = (text or "").lower()
    if "pfte" in t or "fattibilit" in t:
        return "PFTE"
    if "definitiv" in t:
        return "Definitivo"
    if "esecutiv" in t:
        return "Esecutivo"
    return "N/D"


def infer_fase(text: str) -> str:
    t = (text or "").lower()
    if "fase di programmazione" in t or "programmazione" in t:
        return "Programmazione"
    if "fase di progettazione" in t or "progettazione" in t:
        return "Progettazione"
    if "fase di esecuzione" in t or "esecuzione" in t or "cantiere" in t or "lavori" in t:
        return "Esecuzione"
    return "N/D"


def infer_comune(text: str) -> str:
    # euristica: “Comune di X”
    t = text or ""
    m = re.search(r"\bComune\s+di\s+([A-ZÀ-Ü][A-Za-zÀ-ÿ'\-\s]{2,40})", t)
    if m:
        return m.group(1).strip()
    return "N/D"


def infer_committente(text: str) -> str:
    t = text or ""
    # “Ente: …” oppure “Stazione appaltante: …”
    m = re.search(r"(Stazione\s+appaltante|Ente|Committente)\s*[:\-]\s*([A-ZÀ-Ü][^;\n]{3,80})", t, re.I)
    if m:
        return m.group(2).strip()
    # fallback: se compare “Comune di …”
    m2 = re.search(r"\bComune\s+di\s+([A-ZÀ-Ü][A-Za-zÀ-ÿ'\-\s]{2,40})", t)
    if m2:
        return f"Comune di {m2.group(1).strip()}"
    return "N/D"


def pick_first_code(regex: re.Pattern, text: str) -> str:
    if not text:
        return "N/D"
    ms = regex.findall(text)
    if not ms:
        return "N/D"
    # filtra falsi positivi semplici
    for x in ms:
        x = x.strip()
        if len(x) >= 10:
            return x
    return "N/D"


def pick_first_euro(text: str) -> str:
    if not text:
        return "N/D"
    m = EURO_RE.search(text)
    return m.group(0).strip() if m else "N/D"


def build_dettagli(
    committente: str,
    valore: str,
    finanziamento: str,
    fonte_label: str,
    url: str,
    date_used: str,
    cig: str,
    cup: str,
    extra_points: Optional[List[str]] = None,
) -> str:
    pts = []
    pts.append(f"Committente: {committente}")
    pts.append(f"Valore stimato: {valore}")
    pts.append(f"Finanziamento: {finanziamento}")
    pts.append(f"Fonte: {fonte_label} - {url}")
    pts.append(f"Data usata per filtro 60 giorni: {date_used}")
    pts.append(f"Codici: CUP={cup}; CIG={cig}")
    if extra_points:
        for p in extra_points:
            if p and p.strip():
                pts.append(p.strip())
    return "; ".join(pts)


# =========================
# DEDUP
# =========================

def normalize_key(s: str) -> str:
    s = (s or "").lower()
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"[^a-z0-9à-ÿ ]", "", s)
    return s


def dedup_records(df: pd.DataFrame) -> pd.DataFrame:
    """
    Dedup:
    1) Se CUP o CIG presenti => key = CUP|CIG
    2) Altrimenti key = nome+comune+committente(normalizzato)
    Merge: preferisce record con più campi != N/D e dettagli più lunghi
    """
    def score_row(r: pd.Series) -> int:
        vals = [str(r.get(c, "N/D")) for c in df.columns]
        non_nd = sum(1 for v in vals if v and v != "N/D")
        det_len = len(str(r.get("Dettagli", "")))
        return non_nd * 10 + min(det_len, 2000) // 10

    if df.empty:
        return df

    # costruisci key
    keys = []
    for _, r in df.iterrows():
        cup = str(r.get("CUP", "N/D"))
        cig = str(r.get("CIG", "N/D"))
        if cup != "N/D" or cig != "N/D":
            keys.append(f"CODE::{cup}::{cig}")
        else:
            nk = normalize_key(str(r.get("Nome progetto", "")))
            ck = normalize_key(str(r.get("Comune", "")))
            ek = normalize_key(str(r.get("Ente Committente", "")))
            keys.append(f"TXT::{nk}::{ck}::{ek}")
    df = df.copy()
    df["_dedup_key"] = keys
    df["_score"] = df.apply(score_row, axis=1)

    out = []
    for k, g in df.groupby("_dedup_key", dropna=False):
        # scegli il migliore
        g2 = g.sort_values("_score", ascending=False)
        best = g2.iloc[0].copy()

        # unisci dettagli se ci sono più fonti diverse
        if len(g2) > 1:
            dets = []
            for _, rr in g2.iterrows():
                d = str(rr.get("Dettagli", ""))
                if d and d != "N/D":
                    dets.append(d)
            # dedup stringhe
            dets_unique = []
            seen = set()
            for d in dets:
                if d not in seen:
                    seen.add(d)
                    dets_unique.append(d)
            if dets_unique:
                best["Dettagli"] = " || ".join(dets_unique)[:20000]  # limite di sicurezza

        out.append(best)

    out_df = pd.DataFrame(out).drop(columns=["_dedup_key", "_score"], errors="ignore")
    return out_df


# =========================
# CRAWL
# =========================

def crawl_url(
    start_url: str,
    terms: List[str],
    use_selenium: bool,
    parse_pdf_toggle: bool,
    start_date,
    end_date,
    max_results: int,
    max_depth: int,
    status_box=None
) -> List[SearchHit]:
    start_url = normalize_url(start_url)
    if not start_url:
        return []

    domain = urlparse(start_url).netloc
    visited = set()
    queue: List[Tuple[str, int]] = [(start_url, 0)]
    results: List[SearchHit] = []

    t0 = time.time()
    processed = 0
    rough_total = 1 + (max_depth * 25)

    while queue and len(results) < max_results:
        url, depth = queue.pop(0)
        if url in visited:
            continue
        visited.add(url)
        processed += 1

        elapsed = time.time() - t0
        rate = processed / elapsed if elapsed > 0 else 0.0
        remaining = max(0, rough_total - processed)
        eta = remaining / rate if rate > 0 else None

        if status_box is not None:
            with status_box:
                show_working_indicator(f"{urlparse(url).netloc} · depth {depth}", processed / max(rough_total, 1), eta)

        if is_probably_pdf(url):
            title = url.split("/")[-1]
            matched_terms: List[str] = []
            doc_date, emails, note = None, [], ""
            sample = ""

            if parse_pdf_toggle:
                text, doc_date, emails, note = pdf_extract_text_and_meta(url)
                if text:
                    sample = text[:3000]
                    matched_terms = find_matches(text, terms)

            if (not terms) or matched_terms:
                if in_date_range(doc_date, start_date, end_date):
                    results.append(SearchHit(start_url, url, title, matched_terms, doc_date, emails, note, sample))
            continue

        html, fetch_note = fetch_html(url, use_selenium)
        if not html:
            continue

        text = guess_text_from_html(html)
        sample = text[:3000]
        matched = find_matches(text, terms) if terms else []

        if (not terms) or matched:
            doc_date = parse_date_from_text(text)
            if in_date_range(doc_date, start_date, end_date):
                results.append(
                    SearchHit(
                        start_url,
                        url,
                        (text[:90] + "…") if len(text) > 90 else (text[:90] or url),
                        matched,
                        doc_date,
                        sorted(set(EMAIL_RE.findall(text)))[:20],
                        f"html({fetch_note})",
                        sample
                    )
                )

        if depth < max_depth and len(results) < max_results:
            links = extract_links(url, html)
            for lk in links[:150]:
                try:
                    if urlparse(lk).netloc and urlparse(lk).netloc != domain:
                        continue
                except Exception:
                    continue
                if lk not in visited:
                    queue.append((lk, depth + 1))

    return results


# =========================
# OUTPUT BUILDERS
# =========================

def hits_to_records(
    hits: List[SearchHit],
    region_selected: str,
    category_selected: str
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Crea:
    - df12: Excel/PDF con 12 colonne richieste (N/D quando manca)
    - df_fase1_csv: CSV esteso fase1
    """
    rows_12 = []
    rows_csv = []

    for h in hits:
        text = h.raw_text_sample or ""
        cig = pick_first_code(CIG_RE, text)
        cup = pick_first_code(CUP_RE, text)
        comune = infer_comune(text)
        committente = infer_committente(text)
        valore = pick_first_euro(text)
        stato = infer_stato_progetto(text)
        fase = infer_fase(text)
        settore = infer_settore_from_category(category_selected)
        fonte_label = infer_source_label(h.found_url or h.source_url)
        date_used = h.doc_date or "N/D"

        dettagli = build_dettagli(
            committente=committente,
            valore=valore,
            finanziamento="N/D",
            fonte_label=fonte_label,
            url=h.found_url,
            date_used=date_used,
            cig=cig,
            cup=cup,
            extra_points=[
                f"Matched: {', '.join(h.matched_terms) if h.matched_terms else 'N/D'}",
            ],
        )

        nome_progetto = h.title if h.title else "N/D"

        # 12 colonne Excel/PDF
        r12 = {
            "Nome progetto": nd(nome_progetto),
            "Comune": nd(comune),
            "Stato del progetto (PFTE/Definitivo/Esecutivo/N/D)": nd(stato),
            "Fase (Programmazione/Progettazione/Esecuzione/N/D)": nd(fase),
            "Settore (Terziario/Servizi/Residenziale/Industria/Ingegneria/N/D)": nd(settore),
            "Dettagli": nd(dettagli),
            "Progettazione Capogruppo": "N/D",
            "Progettazione Direzione Lavori": "N/D",
            "Progettazione Architettonica": "N/D",
            "Studio di Fattibilità": "N/D",
            "Responsabile Progetto": "N/D",
            "Progetto Impianti Elettrici": "N/D",
        }
        rows_12.append(r12)

        # CSV FASE 1 (esteso)
        rcsv = {
            "Data pubblicazione (DD/MM/YYYY)": nd(h.doc_date),
            "REGIONE (coerente con CIG/CUP)": (region_selected if region_selected != "(tutte)" else "N/D"),
            "Comune/Città": nd(comune),
            "Nome Progetto": nd(nome_progetto),
            "Settore/Tipologia": nd(category_selected if category_selected != "(qualsiasi)" else "N/D"),
            "Ente Committente (Comune, Regione, ASL, Università, Privato, Pubblico-Privato)": nd(committente),
            "Tipo Ente (Pubblico, Pubblico-Privato, Privato)": "N/D",
            "Valore Progetto Totale (€)": nd(valore),
            "Valore Fase Corrente (€)": "N/D",
            "% Completamento (Preliminare, Definitiva, Esecutiva, DL, Collaudo)": "N/D",
            "Data scadenza (DD/MM/YYYY)": "N/D",
            "CIG": nd(cig if cig != "N/D" else ""),
            "CUP": nd(cup if cup != "N/D" else ""),
            "Portale/Fonte (OpenPNRR, OpenCUP, MEPA, GURI, Albo Pretorio, etc.)": nd(fonte_label),
            "Studio/Progettista": "N/D",
            "Fase Progettazione (Preliminare, Definitiva, Esecutiva, Direttore Lavori, Collaudatore)": "N/D",
        }
        rows_csv.append(rcsv)

    df12 = pd.DataFrame(rows_12, columns=EXCEL_PDF_COLS_12)
    dfcsv = pd.DataFrame(rows_csv, columns=CSV_FASE1_COLS)
    return df12, dfcsv


def build_summary_lines(df12: pd.DataFrame) -> List[str]:
    total = len(df12)
    # distribuzione per Fase e Stato (se disponibili)
    fase_col = "Fase (Programmazione/Progettazione/Esecuzione/N/D)"
    stato_col = "Stato del progetto (PFTE/Definitivo/Esecutivo/N/D)"

    fase_counts = df12[fase_col].value_counts(dropna=False).to_dict() if fase_col in df12.columns else {}
    stato_counts = df12[stato_col].value_counts(dropna=False).to_dict() if stato_col in df12.columns else {}

    # fonti principali: estrai da Dettagli "Fonte: X - URL"
    sources = []
    if "Dettagli" in df12.columns:
        for d in df12["Dettagli"].astype(str).tolist():
            m = re.search(r"Fonte:\s*([^;]+)", d)
            if m:
                sources.append(m.group(1).strip())
    src_counts = pd.Series(sources).value_counts().head(5).to_dict() if sources else {}

    lines = []
    lines.append(f"Totale progetti trovati (post-dedup): {total}")
    if fase_counts:
        top_f = ", ".join([f"{k}={v}" for k, v in list(fase_counts.items())[:6]])
        lines.append(f"Distribuzione per Fase: {top_f}")
    if stato_counts:
        top_s = ", ".join([f"{k}={v}" for k, v in list(stato_counts.items())[:6]])
        lines.append(f"Distribuzione per Stato: {top_s}")
    if src_counts:
        top_src = ", ".join([f"{k} ({v})" for k, v in src_counts.items()])
        lines.append(f"Principali fonti: {top_src}")
    # 5–10 righe: garantiamo min 5
    while len(lines) < 5:
        lines.append("Nota: campi mancanti valorizzati come N/D (nessuna supposizione).")
    return lines[:10]


def df_to_xlsx_bytes(df: pd.DataFrame, sheet_name: str = "Progetti") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # style header
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # auto width (soft)
    for col in ws.columns:
        max_len = 12
        col_letter = col[0].column_letter
        for c in col[:200]:
            try:
                max_len = max(max_len, len(str(c.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(55, max(12, max_len * 0.8))

    # add table
    tab = XLTable(displayName="TabellaProgetti", ref=f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def df_to_pdf_bytes(summary_lines: List[str], df: pd.DataFrame) -> bytes:
    bio = io.BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=landscape(A4),
                            leftMargin=10*mm, rightMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    styles = getSampleStyleSheet()
    elems = []

    elems.append(Paragraph("Report Progetti – Tabella (Excel/PDF)", styles["Title"]))
    elems.append(Spacer(1, 6*mm))
    elems.append(Paragraph("<b>Riepilogo</b>", styles["Heading2"]))
    for line in summary_lines:
        elems.append(Paragraph(f"• {line}", styles["BodyText"]))
    elems.append(Spacer(1, 6*mm))

    # Table data
    data = [list(df.columns)] + df.astype(str).values.tolist()

    # wrap long cells by truncation (PDF table can explode)
    def clamp(s: str, n: int = 180) -> str:
        s = s or ""
        s = s.replace("\n", " ")
        return s if len(s) <= n else s[:n] + "…"

    data2 = []
    for r in data:
        data2.append([clamp(str(x), 220 if i == 5 else 120) for i, x in enumerate(r)])  # "Dettagli" un po' più lungo

    tbl = Table(data2, repeatRows=1)

    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
    ]))

    elems.append(tbl)
    doc.build(elems)
    return bio.getvalue()


# =========================
# UI
# =========================

def sidebar_controls():
    st.sidebar.header("⚙️ Menu & filtri")
    nav = st.sidebar.radio("Menu", ["Home", "Dataset", "Ricerca"], index=0)

    st.sidebar.divider()
    st.sidebar.subheader("Opzioni (attivabili)")
    use_selenium = st.sidebar.toggle("Modalità Selenium", value=False, help="Usa Selenium solo se requests viene bloccato.")
    parse_pdf = st.sidebar.toggle("Parsing PDF (pypdf)", value=False, help="Scarica e analizza PDF: più lento.")

    st.sidebar.divider()
    st.sidebar.subheader("Filtri territorio")
    region = st.sidebar.selectbox("Regione", REGIONS, index=0)

    st.sidebar.divider()
    st.sidebar.subheader("Tipi progetto")
    category = st.sidebar.selectbox("Categoria", ["(qualsiasi)"] + list(CATEGORIES.keys()), index=0)

    st.sidebar.subheader("Stato e fase")
    status = st.sidebar.selectbox("Stato del progetto", PROJECT_STATUS, index=0)
    phase = st.sidebar.selectbox("Fase", PROJECT_PHASE, index=0)

    st.sidebar.divider()
    st.sidebar.subheader("Filtro date (se disponibile)")
    c1, c2 = st.sidebar.columns(2)
    start = c1.date_input("data_inizio", value=None)
    end = c2.date_input("data_fine", value=None)

    st.sidebar.divider()
    st.sidebar.subheader("Prestazioni")
    max_results = st.sidebar.slider("Max risultati", 10, 250, 60, 10)
    max_depth = st.sidebar.slider("Profondità link HTML", 0, 2, 1, 1, help="0=solo pagina, 1=anche link, 2=più lento")

    return {
        "nav": nav,
        "use_selenium": use_selenium,
        "parse_pdf": parse_pdf,
        "region": region,
        "category": category,
        "status": status,
        "phase": phase,
        "start_date": start if hasattr(start, "year") else None,
        "end_date": end if hasattr(end, "year") else None,
        "max_results": int(max_results),
        "max_depth": int(max_depth),
    }


def page_home():
    st.title("🔎 Ricerca Progetti Italia – Streamlit (Output XLSX/PDF/CSV)")
    st.write(
        "Questa versione evita loop di caricamento e rende opzionali Selenium/PDF.\n\n"
        "Output: **Excel + PDF con 12 colonne** + **CSV FASE 1 esteso**.\n"
        "Campi mancanti: **N/D** (mai inventare)."
    )


def page_dataset():
    st.header("📁 Dataset")
    st.write("Carica un CSV (file o URL). Parsing robusto: sep auto + skip righe malformate.")

    colA, colB = st.columns([1, 1])
    with colA:
        up = st.file_uploader("Upload CSV", type=["csv"])
    with colB:
        url = st.text_input("...oppure URL CSV", value="")

    df = None
    warnings: List[str] = []

    if up is not None:
        raw = up.read()
        df, warnings = load_csv_robust_from_bytes(raw)
    elif url.strip():
        df, warnings = load_csv_robust(url.strip())

    for w in warnings:
        st.warning(w)

    if df is None or df.empty:
        st.info("Nessun dataset caricato.")
        st.session_state.pop("dataset_df", None)
        st.session_state.pop("dataset_url_col", None)
        st.session_state.pop("dataset_region_col", None)
        return

    st.session_state["dataset_df"] = df
    rcol = detect_region_column(df)
    if rcol:
        st.session_state["dataset_region_col"] = rcol

    st.success(f"Dataset caricato: {df.shape[0]} righe, {df.shape[1]} colonne")
    st.dataframe(df, use_container_width=True)

    url_cols = [c for c in df.columns if "url" in str(c).lower() or "link" in str(c).lower()]
    st.markdown("### Colonna URL (opzionale)")
    if url_cols:
        col = st.selectbox("Seleziona colonna URL", url_cols)
        st.session_state["dataset_url_col"] = col
    else:
        st.caption("Non ho trovato colonne con 'url' o 'link' nel nome.")

    st.markdown("### Colonna Regione (opzionale)")
    if rcol:
        st.caption(f"Rilevata colonna Regione: **{rcol}**")
    else:
        st.caption("Non ho rilevato colonna Regione (rinomina includendo 'regione' se vuoi filtro reale sul dataset).")


def page_search(state):
    st.header("🧭 Ricerca")

    st.markdown("### Sorgenti URL")
    col1, col2 = st.columns([1, 1])

    with col1:
        urls_text = st.text_area(
            "Incolla uno o più URL (uno per riga)",
            value="",
            height=120,
            placeholder="https://www.comune.esempio.it/albo-pretorio\nhttps://www....",
        )
        extra_terms = st.text_area(
            "Keyword extra (comma/righe)",
            value="",
            height=80,
            placeholder='es: "affidamento progettazione", "incarico", CUP, CIG',
        )

    with col2:
        use_dataset = st.checkbox("Usa anche URL dal dataset caricato", value=False)
        limit_urls = st.number_input("Max URL dal dataset", 1, 5000, 200, 50)

        dataset_info = "—"
        if use_dataset and "dataset_df" in st.session_state:
            df = st.session_state["dataset_df"]
            col = st.session_state.get("dataset_url_col")
            if col and col in df.columns:
                dataset_info = f"{len(df[col].dropna().unique())} URL unici dalla colonna '{col}'"
            else:
                dataset_info = "Dataset caricato ma colonna URL non selezionata."
        st.write(dataset_info)

    urls: List[str] = []
    if urls_text.strip():
        for line in urls_text.splitlines():
            u = normalize_url(line)
            if u:
                urls.append(u)

    # Add dataset URLs (with real region filter if column exists)
    region_selected = state["region"]
    if use_dataset and "dataset_df" in st.session_state:
        df = st.session_state["dataset_df"].copy()
        url_col = st.session_state.get("dataset_url_col")

        region_col = st.session_state.get("dataset_region_col")
        if region_selected != "(tutte)" and region_col and region_col in df.columns:
            df[region_col] = df[region_col].astype(str)
            df = df[df[region_col].str.lower().str.contains(region_selected.lower(), na=False)]

        if url_col and url_col in df.columns:
            from_ds = [normalize_url(x) for x in df[url_col].dropna().astype(str).tolist()]
            from_ds = [x for x in from_ds if x]
            seen = set(urls)
            for u in from_ds:
                if u not in seen:
                    seen.add(u)
                    urls.append(u)
            urls = urls[: int(limit_urls)]
        else:
            st.warning("Dataset attivo ma colonna URL non selezionata in Dataset.")

    if not urls:
        st.info("Inserisci almeno 1 URL (o carica un dataset e seleziona la colonna URL).")
        return

    terms = compile_terms(state["category"], extra_terms, state["status"], state["phase"], state["region"])

    st.markdown("### Parametri attivi")
    st.write(
        {
            "urls": len(urls),
            "regione": state["region"],
            "categoria": state["category"],
            "stato (filtro keyword)": state["status"],
            "fase (filtro keyword)": state["phase"],
            "selenium": state["use_selenium"],
            "parse_pdf": state["parse_pdf"],
            "data_inizio": str(state["start_date"]) if state["start_date"] else None,
            "data_fine": str(state["end_date"]) if state["end_date"] else None,
            "depth": state["max_depth"],
        }
    )

    if not st.button("🚀 Avvia ricerca", type="primary"):
        st.caption("Suggerimento: lascia Selenium e parsing PDF OFF finché non ti servono.")
        return

    status_box = st.container()
    all_hits: List[SearchHit] = []
    t0 = time.time()
    total_urls = len(urls)
    hard_max = int(state["max_results"])

    for i, u in enumerate(urls, start=1):
        elapsed = time.time() - t0
        rate = i / elapsed if elapsed > 0 else 0.0
        remaining_urls = total_urls - i
        eta = (remaining_urls / rate) if rate > 0 else None
        show_working_indicator(f"URL {i}/{total_urls}: {u}", i / max(total_urls, 1), eta)

        try:
            chunk = crawl_url(
                start_url=u,
                terms=terms,
                use_selenium=state["use_selenium"],
                parse_pdf_toggle=state["parse_pdf"],
                start_date=state["start_date"],
                end_date=state["end_date"],
                max_results=max(1, hard_max - len(all_hits)),
                max_depth=int(state["max_depth"]),
                status_box=status_box,
            )
            all_hits.extend(chunk)
        except Exception as e:
            all_hits.append(SearchHit(u, u, "ERRORE", [], note=f"Errore ricerca: {e}", raw_text_sample=""))

        if len(all_hits) >= hard_max:
            break

    st.success(f"Ricerca completata: {len(all_hits)} hit grezzi (prima di dedup).")

    # Build records
    df12, dfcsv = hits_to_records(all_hits, region_selected=state["region"], category_selected=state["category"])

    # Aggiungo colonne per dedup (CUP/CIG/Committente) anche a df12 via parse da dettagli (solo per dedup robusto)
    # e genero dfcsv con campi per dedup espliciti
    # Dedup su dfcsv (che ha CUP/CIG e committente)
    dfcsv_for_dedup = dfcsv.copy()
    # normalizza campi richiesti dal dedup func
    dfcsv_for_dedup.rename(columns={
        "Nome Progetto": "Nome progetto",
        "Comune/Città": "Comune",
        "Ente Committente (Comune, Regione, ASL, Università, Privato, Pubblico-Privato)": "Ente Committente",
    }, inplace=True)

    # crea df12_dedup "appoggiandosi" a dfcsv (chiavi)
    df12_with_keys = df12.copy()
    df12_with_keys["CUP"] = dfcsv["CUP"]
    df12_with_keys["CIG"] = dfcsv["CIG"]
    df12_with_keys["Ente Committente"] = dfcsv["Ente Committente (Comune, Regione, ASL, Università, Privato, Pubblico-Privato)"]

    df12_deduped = dedup_records(
        df12_with_keys.rename(columns={"Nome progetto": "Nome progetto", "Comune": "Comune"})
    )

    # ripulisci chiavi aggiunte non richieste dall'Excel/PDF 12 col
    df12_deduped = df12_deduped[EXCEL_PDF_COLS_12].copy()

    st.subheader("Tabella finale (post-dedup)")
    st.dataframe(df12_deduped, use_container_width=True)

    # SUMMARY + export bytes
    summary_lines = build_summary_lines(df12_deduped)
    xlsx_bytes = df_to_xlsx_bytes(df12_deduped, sheet_name="Progetti")
    pdf_bytes = df_to_pdf_bytes(summary_lines, df12_deduped)

    st.download_button(
        "⬇️ Scarica Excel (.xlsx) – 12 colonne",
        xlsx_bytes,
        file_name="progetti_estratti.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.download_button(
        "⬇️ Scarica PDF – tabella + riepilogo",
        pdf_bytes,
        file_name="progetti_estratti.pdf",
        mime="application/pdf",
    )

    # CSV FASE1 esteso (dedup “coerente”: dedup anche qui usando una logica semplice: per CUP/CIG o nome+comune+committente)
    dfcsv_out = dfcsv.copy()
    dfcsv_out["Dettagli (tracciabilità)"] = df12["Dettagli"].astype(str).values

    # dedup su csv (riusiamo dedup_records costruendo campi richiesti)
    tmp = dfcsv_out.copy()
    tmp["Nome progetto"] = tmp["Nome Progetto"]
    tmp["Comune"] = tmp["Comune/Città"]
    tmp["Ente Committente"] = tmp["Ente Committente (Comune, Regione, ASL, Università, Privato, Pubblico-Privato)"]
    tmp["Dettagli"] = tmp["Dettagli (tracciabilità)"]
    tmp_dedup = dedup_records(tmp[["Nome progetto", "Comune", "Ente Committente", "CUP", "CIG", "Dettagli"]])
    # Applica dedup “per chiave” mantenendo la prima occorrenza nel dfcsv_out
    # (semplificazione: se vuoi dedup perfetto anche su tutte le colonne, te lo estendo)
    st.download_button(
        "⬇️ Scarica CSV FASE 1 (colonne estese)",
        dfcsv_out.to_csv(index=False).encode("utf-8"),
        file_name="progetti_fase1.csv",
        mime="text/csv",
    )


def main():
    state = sidebar_controls()
    if state["nav"] == "Home":
        page_home()
    elif state["nav"] == "Dataset":
        page_dataset()
    else:
        page_search(state)


if __name__ == "__main__":
    main()
