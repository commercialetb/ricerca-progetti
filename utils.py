# utils.py
from __future__ import annotations

import re
from io import BytesIO
from typing import Dict, List, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def normalize_progettista(nome_raw: str) -> str:
    """Normalizza: 'Arch. Rossi Mario' → 'ROSSI MARIO'"""
    if not nome_raw:
        return "NON TROVATO"

    s = str(nome_raw).strip()

    titoli = ["arch", "ing", "dott", "prof", "studio", "geom", "per", "p.i."]
    for t in titoli:
        s = re.sub(rf"\b{re.escape(t)}\.?\s*", "", s, flags=re.I)

    s = re.sub(r"\s+", " ", s).strip()

    parts = s.upper().split()
    if len(parts) >= 2:
        return f"{parts[-2]} {parts[-1]}"
    return s.upper()


def _get(d: Dict[str, Any], *keys: str, default=None):
    for k in keys:
        if k in d and d[k] not in (None, ""):
            return d[k]
    return default


def create_csv_segments(leads: List[Dict]) -> str:
    """Ritorna un CSV unico con colonna 'segment' (A/B/C/D)."""
    rows = []
    for lead in leads:
        score = _get(lead, "lead_quality_score", "LEAD_SCORE", default=0) or 0
        try:
            score_f = float(score)
        except Exception:
            score_f = 0.0

        if score_f >= 8.5:
            segment = "A"
        elif score_f >= 7.0:
            segment = "B"
        elif score_f >= 5.0:
            segment = "C"
        else:
            segment = "D"

        rows.append(
            {
                "segment": segment,
                "progettista_norm": _get(lead, "progettista_norm", "PROGETTISTA", default="NON TROVATO"),
                "email": _get(lead, "email", "EMAIL", default="non trovato"),
                "telefono": _get(lead, "telefono", "TELEFONO", default="non trovato"),
                "validation_score": _get(lead, "validation_score", "VALIDATION_SCORE", default=0),
                "lead_quality_score": score_f,
                "comune": _get(lead, "comune", "COMUNE", default=""),
                "regione": _get(lead, "regione", "REGIONE", default=""),
                "pdf_source": _get(lead, "pdf_source", "PDF_SOURCE", default=""),
            }
        )

    df = pd.DataFrame(rows)
    return df.to_csv(index=False)


def create_excel_4sheets(leads: List[Dict]) -> bytes:
    """Excel con 4 sheet: CONTACT_LIST, PROJECTS_BY_DESIGNER, QUALITY_METRICS, OUTREACH_SEGMENTS."""
    wb = Workbook()

    # CONTACT_LIST
    ws1 = wb.active
    ws1.title = "CONTACT_LIST"
    contact_rows = []
    for lead in leads:
        contact_rows.append(
            {
                "PROGETTISTA": _get(lead, "progettista_norm", "PROGETTISTA", default="NON TROVATO"),
                "EMAIL": _get(lead, "email", "EMAIL", default="non trovato"),
                "TELEFONO": _get(lead, "telefono", "TELEFONO", default="non trovato"),
                "VALIDATION_SCORE": _get(lead, "validation_score", "VALIDATION_SCORE", default=0),
                "LINKEDIN": _get(lead, "linkedin", "LINKEDIN", default=""),
            }
        )
    df_contacts = pd.DataFrame(contact_rows)
    for r in dataframe_to_rows(df_contacts, index=False, header=True):
        ws1.append(r)

    # PROJECTS_BY_DESIGNER
    ws2 = wb.create_sheet("PROJECTS_BY_DESIGNER")
    proj_rows = []
    for lead in leads:
        proj_rows.append(
            {
                "PROGETTISTA": _get(lead, "progettista_norm", "PROGETTISTA", default="NON TROVATO"),
                "COMUNE": _get(lead, "comune", "COMUNE", default=""),
                "PROVINCIA": _get(lead, "provincia", "PROVINCIA", default=""),
                "REGIONE": _get(lead, "regione", "REGIONE", default=""),
                "DATA_DELIBERA": _get(lead, "data_delibera", default=""),
                "CUP": _get(lead, "cup", default=""),
                "CIG": _get(lead, "cig", default=""),
                "IMPORTO": _get(lead, "importo", default=""),
                "PDF_SOURCE": _get(lead, "pdf_source", default=""),
                "PORTAL_URL": _get(lead, "portal_url", default=""),
            }
        )
    df_proj = pd.DataFrame(proj_rows)
    for r in dataframe_to_rows(df_proj, index=False, header=True):
        ws2.append(r)

    # QUALITY_METRICS
    ws3 = wb.create_sheet("QUALITY_METRICS")
    qm_rows = []
    for lead in leads:
        fonti = _get(lead, "fonti", default=[])
        qm_rows.append(
            {
                "PROGETTISTA": _get(lead, "progettista_norm", "PROGETTISTA", default="NON TROVATO"),
                "VALIDATION_SCORE": _get(lead, "validation_score", "VALIDATION_SCORE", default=0),
                "LEAD_QUALITY_SCORE": _get(lead, "lead_quality_score", "LEAD_SCORE", default=0),
                "FONTI": ", ".join(fonti) if isinstance(fonti, list) else str(fonti),
                "ERROR_AI": _get(lead, "error_ai", default=""),
                "ERROR": _get(lead, "error", default=""),
            }
        )
    df_qm = pd.DataFrame(qm_rows)
    for r in dataframe_to_rows(df_qm, index=False, header=True):
        ws3.append(r)

    # OUTREACH_SEGMENTS
    ws4 = wb.create_sheet("OUTREACH_SEGMENTS")
    df_seg = pd.read_csv(BytesIO(create_csv_segments(leads).encode("utf-8")))
    for r in dataframe_to_rows(df_seg, index=False, header=True):
        ws4.append(r)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def generate_outreach_templates(top_leads: List[Dict]) -> str:
    """Genera email personalizzate (testo)."""
    templates = []
    for lead in top_leads:
        proj = _get(lead, "comune", "COMUNE", default="(comune)")
        name = _get(lead, "progettista_norm", "PROGETTISTA", default="(progettista)")
        cup = _get(lead, "cup", default="non trovato")
        cig = _get(lead, "cig", default="non trovato")

        template = (
            f"Subject: Complimenti per il progetto a {proj}\n\n"
            f"Gentile {name},\n\n"
            f"ho visto un vostro intervento a {proj} (CUP: {cup}, CIG: {cig}).\n"
            f"Mi farebbe piacere capire se state lavorando su nuove commesse nel 2026: possiamo sentirci 10 minuti?\n\n"
            f"Cordiali saluti,\n"
            f"[Tuo Nome]\n"
        )
        templates.append(template)
    return "\n\n---\n\n".join(templates)
